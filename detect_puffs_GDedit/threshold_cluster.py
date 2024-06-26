# -*- coding: utf-8 -*-
"""
Created on Tue Sep 02 16:54:20 2014
@author: Kyle Ellefsen

The logic of the puff detection algorithm works like this.  We start with three movies:
1) F/F0 Movie
2) Normalized Movie
3) Blurred Movie

These get passed into the threshold_cluster() function, which creates the algorithm_gui.
The algorithm_gui waits for the 'Set Threshold' button to be pressed, and then calls the getClusters function.
The getClusters() function calls getHigherPoints.  It then creates the clusters object, which in turn creates the ClusterViewBox which displays the higher_pts and can be drawn on.
Once a set of the higher_pts have been circled, the clusters object calls its manuallySelectClusterCenters and makes the cluster_im window, which displays the colored clusters.
Once the clusters are created and thresholded by size, pressing the button 'Fit Gaussians' tells the clusters object to create a Puffs object and fit each cluster to a gaussian function.
These puffs are displayed on the F/F0 movie as points.


"""
import pickle
import bz2
import ntpath
import itertools
from distutils.version import StrictVersion
import os, inspect,sys
import numpy as np
from qtpy import QtWidgets, QtGui, QtCore
from qtpy.QtWidgets import *
from qtpy.QtGui import *
from qtpy.QtCore import *
from qtpy.QtWidgets import qApp
from qtpy import uic
import pyqtgraph as pg
from scipy import ndimage
import pyqtgraph.opengl as gl
import scipy.stats as st
from scipy import spatial
import matplotlib
from pyqtgraph.dockarea import *


import flika
try:
    flika_version = flika.__version__
except AttributeError:
    flika_version = '0.0.0'
if StrictVersion(flika_version) < StrictVersion('0.1.0'):
    import global_vars as g
    from process.BaseProcess import BaseProcess, WindowSelector, SliderLabel, CheckBox
    from window import Window
    from roi import ROI_rectangle, makeROI
    from process.file_ import open_file_gui, open_file
    g.alert('The detect_puffs no longer works with older versions of Filka. Delete your current version of Flika and reinstall the newer version before using this plugin. ')
else:
    from flika import global_vars as g
    from flika.window import Window
    from flika.roi import ROI_rectangle, makeROI
    from flika.process.file_ import open_file
    from flika.utils.misc import open_file_gui
    from flika.process import *
    from flika.utils.misc import save_file_gui
    if StrictVersion(flika_version) < StrictVersion('0.2.23'):
        from flika.process.BaseProcess import BaseProcess, WindowSelector, SliderLabel, CheckBox
    else:
        from flika.utils.BaseProcess import BaseProcess, WindowSelector, SliderLabel, CheckBox

from .puffs import Puffs, Puff
from .groups import GroupAnalyzer, Group, Groups
from .higher_pts import getHigherPoints
from .clusters import Clusters, ClusterViewBox, ROI

if StrictVersion(flika_version) < StrictVersion('0.1.0'):
    flika_icon = QIcon('images/favicon.png')
    threshold_cluster_ui = os.path.join(os.getcwd(), 'plugins', 'detect_puffs', 'threshold_cluster.ui')
else:
    flika_icon = QIcon('flika/images/favicon.png')
    threshold_cluster_ui = os.path.join(os.path.dirname(__file__), 'threshold_cluster.ui')



class OddSlider(SliderLabel):
    def __init__(self,demicals=0):
        SliderLabel.__init__(self,demicals)
    def updateSlider(self,value):
        if value%2==0:
            if value<self.slider.value():
                value-=1
            else:
                value+=1
            self.label.setValue(value)
        self.slider.setValue(int(value*10**self.decimals))
    def updateLabel(self,value):
        if value%2==0:
            value-=1
        self.label.setValue(value)


def launch_docs():
    url = 'https://github.com/kyleellefsen/detect_puffs'
    QDesktopServices.openUrl(QUrl(url))


def load_flika_file_gui():
    prompt = 'Open .flika file'
    filetypes = '*.flika'
    filename = open_file_gui(prompt=prompt, filetypes=filetypes)
    load_flika_file(filename)


def load_flika_file(filename=None):
    print('loading ... {}'.format(filename))
    with bz2.BZ2File(filename, 'rb') as f:
        persistentInfo=pickle.load(f)
    data_path=os.path.dirname(filename)
    data_filename=ntpath.basename(os.path.splitext(filename)[0])
    data_filename_matching=[name for name in os.listdir(data_path) if data_filename==os.path.splitext(name)[0] and os.path.splitext(name)[1] in ['.tif', '.stk', '.tiff']]
    if len(data_filename_matching)==0:
        msg='To load the .flika file, your original data file must be in the same directory and named with the same name as your flika file.  Original file not found'
        g.m.statusBar().showMessage(msg); print(msg)
        return None
    data_filename=data_filename_matching[0]
    raw_data_window=open_file(os.path.join(data_path,data_filename))
    
    if hasattr(persistentInfo,'data_window_commands'):
        commands=persistentInfo.data_window_commands
        for command in commands[1:]:
            exec(command)
        data_window=g.currentWindow
        data_window.setName('Data Window (F/F0)')
    else:
        data_window=raw_data_window
        data_window.setName('Data Window (raw)')
    puffAnalyzer=PuffAnalyzer(data_window,None,None,None,persistentInfo)
    g.windows.append(puffAnalyzer)


class Threshold_cluster(BaseProcess):
    """threshold_cluster(data_window, normalized_window, blurred_window, roi_width, paddingXY, paddingT_pre, paddingT_post, maxSigmaForGaussianFit, rotatedfit)
    Performs lots of analyses on puffs
    
    Parameters:
        | data_window (Window) --  This Window is the F/F0 window
        | normalized_window (Window) -- This Window should have the baseline at 0.  It will be used for gaussian fitting and kinetics.
        | blurred_window (Window) -- Usually the gaussian blurred version of the normalized_window.
        | roi_width (int) -- The width of the ROI in pixels
        | paddingXY (int) -- How many pixels do you want to pad the ROI by when fitting with gaussian
        | paddingT_pre (int) -- How many frames before the event detection should we look for the start of the event.
        | paddingT_post (int)  -- How many frames after the event detection should we look for the end of the event.
        | maxSigmaForGaussianFit (int) -- When fitting with a gaussian, what is the upper bound for the sigma (or sigmax and sigmay) parameter 
        | rotatedfit (bool) -- Set this to True to fit to a 2d rotating gaussian.
        | radius (float) -- Puffs seperated by less than this amount (in pixels) will be grouped together in a group.
        | maxPuffLen (int) -- maximum possible duration of puff length in frames.  If puffs last longer, use a bigger value.  This affects 'chunking' when determining distances between pixels.  It's better to overestimate, although increasing too much will slow down analysis.
        | maxPuffDiameter (int) -- This affects 'chunking' when determining distances between pixels.    This value divided by two is also used as the radius for calculating densities
        | blur_thresh (float) -- Get rid of all the points you are sure don't belong to puffs.
        | time_factor (float) -- When looking for the nearest point of higher density, time and space don't need to be weighted equally.  This is how much I will 'shrink' time in order to calculate distances relative to pixels.  Practically, the larger this number, the more points separated in time will tend to be clustered together.  Use a higher time factor if you are sampling at high frame rates.  
        | load_flika_file (bool) -- If this is checked, the flika file with the same name as the data_window file will be loaded, if it exists.
    Returns:
        newWindow
    """
    def __init__(self):
        super().__init__()

    def get_init_settings_dict(self):
        s = dict()
        s['roi_width'] = 1
        s['paddingXY'] = 20
        s['paddingT_pre'] = 15
        s['paddingT_post'] = 15
        s['maxSigmaForGaussianFit'] = 10
        s['rotatedfit'] = True
        s['radius'] = np.sqrt(2)
        s['maxPuffLen'] = 15
        s['maxPuffDiameter'] = 10
        s['blur_thresh'] = 2.7
        s['time_factor'] = 1
        s['load_flika_file'] = True
        return s

    def gui(self):
        self.gui_reset()
        self.data_window=WindowSelector()
        self.normalized_window=WindowSelector()
        self.blurred_window=WindowSelector()
        roi_width=OddSlider(0)
        roi_width.setRange(1,21)
        roi_width.setSingleStep(2)
        paddingXY=SliderLabel(0)
        paddingXY.setRange(0,1000)
        paddingT_pre=SliderLabel(0); paddingT_pre.setRange(0,1000)
        paddingT_post=SliderLabel(0); paddingT_post.setRange(0,1000)
        maxSigmaForGaussianFit=SliderLabel(0); maxSigmaForGaussianFit.setRange(0,100)
        rotatedfit=CheckBox()
        radius=SliderLabel(2)
        radius.setRange(.01,10)
        radius.setSingleStep(.1)
        maxPuffLen=SliderLabel(0)
        maxPuffLen.setRange(1,1000)
        maxPuffDiameter=SliderLabel(0)
        maxPuffDiameter.setRange(1,100)
        blur_thresh=SliderLabel(2)
        blur_thresh.setRange(0,6)
        time_factor=SliderLabel(3)
        time_factor.setRange(0,20)
        load_flika_file=CheckBox()
        
        self.autoInput_win = AutomatedInput_win()
        self.automateInput=CheckBox()
        self.automateInput.stateChanged.connect(self.automateInput_dialog)  
            
       
        self.items.append({'name':'automated_Input','string':'Automate input', 'object': self.automateInput})  
        self.items.append({'name':'data_window','string':'Data window containing F/F0 data', 'object': self.data_window})
        self.items.append({'name':'normalized_window','string':'Normalized window containing data with baseline at 0','object': self.normalized_window})
        self.items.append({'name':'blurred_window','string': 'Gaussian Blurred normalized window','object': self.blurred_window})       
        self.items.append({'name':'roi_width','string':'roi_width','object':roi_width})
        self.items.append({'name':'paddingXY','string':'paddingXY','object':paddingXY})
        self.items.append({'name':'paddingT_pre','string':'paddingT_pre','object':paddingT_pre})
        self.items.append({'name':'paddingT_post','string':'paddingT_post','object':paddingT_post})
        self.items.append({'name':'maxSigmaForGaussianFit','string':'maxSigmaForGaussianFit','object':maxSigmaForGaussianFit})
        self.items.append({'name':'rotatedfit','string':'rotatedfit','object':rotatedfit})
        self.items.append({'name':'radius','string':'radius','object':radius})
        self.items.append({'name':'maxPuffLen','string':'maxPuffLen','object':maxPuffLen})
        self.items.append({'name':'maxPuffDiameter','string':'maxPuffDiameter','object':maxPuffDiameter})
        self.items.append({'name':'blur_thresh','string':'blur_thresh','object':blur_thresh})
        self.items.append({'name':'time_factor','string':'time_factor','object':time_factor})
        self.items.append({'name':'load_flika_file','string':'load_flika_file','object':load_flika_file})
        super().gui()
        self.ui.setGeometry(QRect(676, 55, 1231, 822))

    def __call__(self, data_window, normalized_window, blurred_window, roi_width=3, paddingXY=20, paddingT_pre=15,
                 paddingT_post=15, maxSigmaForGaussianFit=10, rotatedfit=True, radius=np.sqrt(2), maxPuffLen=15,
                 maxPuffDiameter=10, blur_thresh=1, time_factor=1, load_flika_file=True, keepSourceWindow=False):
        g.m.statusBar().showMessage('Performing {}...'.format(self.__name__))
        
        try:
            self.autoInput_win.close()
        except:
            pass
        
        filename = data_window.filename
        filename = os.path.splitext(filename)[0]+'.flika'
        if load_flika_file and os.path.isfile(filename):
            print('Found persistentInfo file.  Loading previous results')
            with bz2.BZ2File(filename, 'rb') as f:
                persistentInfo = pickle.load(f)
            puffAnalyzer = PuffAnalyzer(data_window, normalized_window, blurred_window, None, persistentInfo)
        else:
            udc=dict()
            udc['roi_width'] = roi_width
            udc['paddingXY'] = paddingXY
            udc['paddingT_pre'] = paddingT_pre
            udc['paddingT_post'] = paddingT_post
            udc['maxSigmaForGaussianFit'] = maxSigmaForGaussianFit
            udc['rotatedfit'] = rotatedfit
            udc['radius'] = radius  # radius -- all puffs this distance away (measured in pixels) from each other will automatically be grouped together into a group
            udc['maxPuffLen'] = maxPuffLen
            udc['maxPuffDiameter'] = maxPuffDiameter
            if blur_thresh is None:
                if 'threshold_cluster' in g.settings['baseprocesses'].keys():
                    blur_thresh = g.settings['baseprocesses']['threshold_cluster']['blur_thresh']
                else:
                    blur_thresh = 1
            udc['blur_thresh'] = blur_thresh
            udc['time_factor'] = time_factor
            puffAnalyzer = PuffAnalyzer(data_window, normalized_window, blurred_window, udc)
        
        
        g.windows.append(puffAnalyzer)
        g.m.statusBar().showMessage('Finished with {}.'.format(self.__name__))
        return puffAnalyzer


    def automateInput_dialog(self):
        if self.automateInput.isChecked():
            self.autoInput_win.show()
        else:
            self.autoInput_win.close()
            

threshold_cluster = Threshold_cluster()


class AutomatedInput_win(QtWidgets.QDialog):
    def __init__(self, parent = None):
        super(AutomatedInput_win, self).__init__(parent)

        self.rawData_window=WindowSelector()
        self.rawData_window.valueChanged.connect(self.updateSliders)
            
        self.blackLevel=SliderLabel(1)
        self.blackLevel.setRange(0,10000)
        self.blackLevel.setSingleStep(.1)
        self.blackLevel.setValue(0)   
        
        self.firstFrame=SliderLabel(0)
        self.firstFrame.setRange(0,1000)
        self.firstFrame.setSingleStep(1)        
        
        self.nFrames=SliderLabel(0)
        self.nFrames.setRange(1,1000)
        self.nFrames.setSingleStep(1) 
        self.nFrames.setValue(30)        
        
        self.blur=SliderLabel(0)
        self.blur.setRange(0,100)
        self.blur.setSingleStep(1)    
        self.blur.setValue(2)

        self.trimRaw=CheckBox()

        self.firstFrame_trim=SliderLabel(0)
        self.firstFrame_trim.setRange(0,10000)
        self.firstFrame_trim.setSingleStep(1) 
        self.firstFrame_trim.setValue(0)
        
        self.lastFrame_trim=SliderLabel(0)
        self.lastFrame_trim.setRange(1,100000)
        self.lastFrame_trim.setSingleStep(1)

        self.firstFrame_flash=SliderLabel(0)
        self.firstFrame_flash.setRange(0,10000)
        self.firstFrame_flash.setSingleStep(1) 
        
        self.lastFrame_flash=SliderLabel(0)
        self.lastFrame_flash.setRange(1,100000)
        self.lastFrame_flash.setSingleStep(1)

        self.removeFlash=CheckBox()    

        self.nFrame_butter=SliderLabel(0)
        self.nFrame_butter.setRange(0,1000)
        self.nFrame_butter.setSingleStep(1)  
        self.nFrame_butter.setValue(25)        
        
        self.filter_order = SliderLabel(0)
        self.filter_order.setRange(1,10)
        self.filter_order.setSingleStep(1)
        self.filter_order.setValue(1)
               
        self.low = SliderLabel(2)
        self.low.setRange(0,1000)
        self.low.setSingleStep(0.01) 
        self.low.setValue(0.1)         
        
        self.high = SliderLabel(2)
        self.high.setRange(0,1000)
        self.high.setSingleStep(0.01)        
        self.high.setValue(1)         
        
        self.framerate = SliderLabel(0)
        self.framerate.setRange(1,5000)
        self.framerate.setSingleStep(1)
        self.framerate.setValue(227)      
        
        self.saveDetails=CheckBox()

        self.run = QtWidgets.QPushButton('Create input windows')
        self.run.clicked.connect(self.automateInput)
        
        #window geometry
        self.left = 300
        self.top = 300
        self.width = 500
        self.height = 500

        #labels
        self.raw_data_window_label = QtWidgets.QLabel('raw data window')
        self.firstFrame_trim_label = QtWidgets.QLabel('trim raw: first Frame')          
        self.lastFrame_trim_label = QtWidgets.QLabel('trim raw: last Frame')        
        self.trimRaw_label = QtWidgets.QLabel('trim raw data')                
        self.blackLevel_label = QtWidgets.QLabel('camera black level to subtract')         
        self.firstFrame_label = QtWidgets.QLabel('DF/F0 ratio: firstFrame')          
        self.nFrames_label = QtWidgets.QLabel('DF/F0 ratio: nFrames')          
        self.firstFrame_flash_label = QtWidgets.QLabel('flash: first Frame')          
        self.lastFrame_flash_label = QtWidgets.QLabel('flash: last Frame')         
        self.removeFlash_label = QtWidgets.QLabel('remove flash artifact')                 
        self.nFrame_butter_label = QtWidgets.QLabel('butter filter: n frames to trim from ends')                                  
        self.filter_order_label = QtWidgets.QLabel('butterworth filter: order')
        self.low_label = QtWidgets.QLabel('butterworth filter: low')
        self.high_label = QtWidgets.QLabel('butterworth filter: high')
        self.framerate_label = QtWidgets.QLabel('butterworth filter: framerate')    
        self.blur_label = QtWidgets.QLabel('Gaussian blur: sigma')
        self.saveDetails_label = QtWidgets.QLabel('Save details:')        

                           
        #grid layout
        layout = QtWidgets.QGridLayout()
        layout.setSpacing(3)
        #widgets
        layout.addWidget(self.rawData_window, 1, 1)
        layout.addWidget(self.firstFrame_trim, 2, 1)          
        layout.addWidget(self.lastFrame_trim, 3, 1)       
        layout.addWidget(self.trimRaw, 4, 1)                
        layout.addWidget(self.blackLevel, 6, 1)        
        layout.addWidget(self.firstFrame, 7, 1)         
        layout.addWidget(self.nFrames, 8, 1)         
        layout.addWidget(self.firstFrame_flash, 9, 1)          
        layout.addWidget(self.lastFrame_flash, 10, 1)         
        layout.addWidget(self.removeFlash, 11, 1)                 
        layout.addWidget(self.nFrame_butter, 12, 1)                                   
        layout.addWidget(self.filter_order, 13, 1) 
        layout.addWidget(self.low, 14, 1)
        layout.addWidget(self.high, 15, 1) 
        layout.addWidget(self.framerate, 16, 1)       
        layout.addWidget(self.blur, 17, 1)    
        layout.addWidget(self.run, 19, 1)   
        layout.addWidget(self.saveDetails, 18, 1)           
        #labels
        layout.addWidget(self.raw_data_window_label, 1, 0)
        layout.addWidget(self.firstFrame_trim_label, 2, 0)          
        layout.addWidget(self.lastFrame_trim_label, 3, 0)       
        layout.addWidget(self.trimRaw_label, 4, 0)                
        layout.addWidget(self.blackLevel_label, 6, 0)        
        layout.addWidget(self.firstFrame_label, 7, 0)         
        layout.addWidget(self.nFrames_label, 8, 0)         
        layout.addWidget(self.firstFrame_flash_label, 9, 0)          
        layout.addWidget(self.lastFrame_flash_label, 10, 0)         
        layout.addWidget(self.removeFlash_label, 11, 0)                 
        layout.addWidget(self.nFrame_butter_label, 12, 0)                                  
        layout.addWidget(self.filter_order_label, 13, 0) 
        layout.addWidget(self.low_label, 14, 0)
        layout.addWidget(self.high_label, 15, 0) 
        layout.addWidget(self.framerate_label, 16, 0)       
        layout.addWidget(self.blur_label, 17, 0) 
        layout.addWidget(self.saveDetails_label, 18, 0)   
        
        self.setLayout(layout)
        self.setGeometry(self.left, self.top, self.width, self.height)

        #add window title
        self.setWindowTitle("Automated Input Options")
        
 
    def updateSliders(self):
        nFrames = self.rawData_window.value().image.shape[0]
        self.firstFrame.setRange(0,nFrames)
        self.nFrames.setRange(1,nFrames)        
        self.firstFrame_trim.setRange(0,nFrames)
        self.lastFrame_trim.setRange(1,nFrames)
        self.firstFrame_flash.setRange(0,nFrames)
        self.lastFrame_flash.setRange(1,nFrames)
        self.nFrame_butter.setRange(0,int(nFrames/2))
 
    def automateInput(self):
        #initiate save details list
        self.saveDetailsList = []
        print('automated input activated')
        rawData = Window(self.rawData_window.value().image)
        rawData.setAsCurrentWindow()
        self.windowName = self.rawData_window.value().filename
        self.saveDetailsList.append('Raw Data: {}'.format(self.windowName))

        #trim frames
        if self.trimRaw.isChecked():
            rawData = trim(self.firstFrame_trim.value(), self.lastFrame_trim.value(), increment=1, delete=False, keepSourceWindow=False)
            self.saveDetailsList.append('Trimming: trim({}, {}, increment=1, delete=False, keepSourceWindow=False)'.format(self.firstFrame_trim.value(), self.lastFrame_trim.value()))
                
        #create DF/F0
        data_window = ratio(self.firstFrame.value(), self.nFrames.value(), 'average', self.blackLevel.value())  # ratio(first_frame, nFrames, ratio_type). Now we are in F/F0
        data_window.setName('Data Window (F/F0)')
        self.saveDetailsList.append('DF/F0 data window: ratio({}, {}, {}, {})'.format(self.firstFrame.value(), self.nFrames.value(), "average", self.blackLevel.value()))

        #remove flash artifact
        if self.removeFlash.isChecked():
            data_window = set_value(1,self.firstFrame_flash.value(),self.lastFrame_flash.value(), restrictToROI=False, keepSourceWindow=False)
            self.saveDetailsList.append('Remove flash: set_value(1,{}, {}, restrictToROI=False, KeepSourceWindow=False)'.format(self.firstFrame_flash.value(),self.lastFrame_flash.value()))
            data_window.setName('Data Window (F/F0)')

        #create normalised image using butterworth filtered image
        butter_image = butterworth_filter(self.filter_order.value(), self.low.value(), self.high.value(), self.framerate.value(), keepSourceWindow=True)            
        self.saveDetailsList.append('Butterworth filter: butterworth_filter({}, {}, {}, {}, keepSourceWindow=True)'.format(self.filter_order.value(), self.low.value(), self.high.value(), self.framerate.value()))

        #get n frames
        nFrames = butter_image.image.shape[0]            

        #trim ends of filtered image to remove windowing artifacts
        trim(self.nFrame_butter.value(), nFrames-self.nFrame_butter.value(), increment=1, keepSourceWindow=False)
        self.saveDetailsList.append('Trim ends: trim({}, {}, increment=1, keepSourceWindow=False)'.format(self.nFrame_butter.value(), nFrames-self.nFrame_butter.value()))

        #ratio filtered image by standard deviation
        norm_window = ratio(self.firstFrame.value()+self.nFrame_butter.value(), self.nFrames.value(), 'standard deviation', keepSourceWindow=False)
        self.saveDetailsList.append('Normalized Window: ratio({}, {}, "standard deviation", keepSourceWindow=False)'.format(self.firstFrame.value()+self.nFrame_butter.value(), self.nFrames.value()))

        #set norm_window 
        norm_window.setName('Normalized Window')            

        #create blurred image
        blurred_window = gaussian_blur(self.blur.value(), norm_edges=True, keepSourceWindow=True)
        blurred_window.setName('Blurred Window') 
        self.saveDetailsList.append('Blurred Window: gaussian_blur({}, norm_edges=True, keepSourceWindow=True)'.format(self.blur.value()))

        #set DF/F0 window as current
        data_window.setAsCurrentWindow()

        #trimDF/F0 window to match norm and blurred
        data_window = trim(self.nFrame_butter.value(), nFrames-self.nFrame_butter.value(), increment=1, keepSourceWindow=False)   
        self.saveDetailsList.append('Trim DF/F0 ends: trim({}, {}, increment=1, keepSourceWindow=False)'.format(self.nFrame_butter.value(), nFrames-self.nFrame_butter.value()))

        #rename DF/F0
        data_window.setName('Data Window (F/F0)')

        #update window selectors
        threshold_cluster.data_window.setWindow(window=data_window)
        threshold_cluster.normalized_window.setWindow(window=norm_window)
        threshold_cluster.blurred_window.setWindow(window=blurred_window)

        #save details
        if self.saveDetails.isChecked():
            
            filepath = self.windowName.split('.')[0] + '_autoInput.txt'
            
            with open(filepath, 'w') as file_handler:
                for item in self.saveDetailsList:
                    file_handler.write("{}\n".format(item))

        #finised message
        g.m.statusBar().showMessage('Generated Input Windows')


class PuffAnalyzer(QWidget):
    def __init__(self, data_window, normalized_window, blurred_window, udc, persistentInfo=None, parent=None):
        """ udc -- all the user defined constants """
        super(PuffAnalyzer,self).__init__(parent)  # Create window with ImageView widget
        g.puffAnalyzer = self
        self.name = 'Puff Analyzer'
        self.setWindowIcon(flika_icon)
        self.data_window = data_window
        self.normalized_window = normalized_window
        self.blurred_window = blurred_window
        self.udc = udc
        self.mt, self.mx, self.my = self.data_window.image.shape
        self.puffs = None
        self.l = None
        self.trash = None
        self.clusters = None
        self.groups = None
        self.gettingClusters = False
        self.generatingClusterMovie = False
        self.algorithm_gui = uic.loadUi(threshold_cluster_ui)
        self.algorithm_gui.setWindowIcon(flika_icon)
        self.algorithm_gui.show()
        if persistentInfo is not None:
            if 'roi_width' not in persistentInfo.udc.keys():
                persistentInfo.udc['roi_width'] = 3
            self.loadPersistentInfo(persistentInfo)
        else:
            self.algorithm_gui.density_movie_layout.addWidget(self.blurred_window)
            self.algorithm_gui.thresh_slider.setValue(udc['blur_thresh'])
            self.algorithm_gui.paddingXY.setText(str(udc['paddingXY']))
            self.algorithm_gui.maxSigmaForGaussianFit.setText(str(udc['maxSigmaForGaussianFit']))
            self.algorithm_gui.rotatedfit.setText(str(udc['rotatedfit']))
            self.algorithm_gui.thresh_button1.pressed.connect(self.getClusters)

    def getClusters(self):
        if self.gettingClusters:
            return None
        self.gettingClusters = True
        self.udc['blur_thresh'] = self.algorithm_gui.thresh_slider.value()
        blurred = self.blurred_window.image
        higher_pts, idxs = getHigherPoints(blurred, self.udc)
        self.algorithm_gui.tabWidget.setCurrentIndex(1)
        qApp.processEvents()
        self.clusters = Clusters(higher_pts, idxs, blurred.shape, self)
        self.algorithm_gui.tabWidget.setCurrentIndex(1)
        qApp.processEvents()
        self.gettingClusters = False

    def loadPersistentInfo(self, persistentInfo):
        self.udc = persistentInfo.udc
        self.clusters = Clusters(None, None, None, self, persistentInfo)
        self.groups = Groups(self)
        if hasattr(persistentInfo,'groups'):
            for i in np.arange(len(persistentInfo.groups)):
                self.groups.append(Group([puff for puff in g.puffAnalyzer.puffs.puffs if puff.starting_idx in persistentInfo.groups[i]]))
        else:  # 'groups' used to be named 'sites'.  This is for legacy .flika files.
            for i in np.arange(len(persistentInfo.sites)):
                self.groups.append(Group([puff for puff in g.puffAnalyzer.puffs.puffs if puff.starting_idx in persistentInfo.sites[i]]))
        self.trash = Trash(self)
        for i in np.arange(len(persistentInfo.puffs)):
            if persistentInfo.puffs[i]['trashed']:
                puff = [puff for puff in g.puffAnalyzer.puffs.puffs if puff.starting_idx == i][0]
                self.trash.append(puff)
                self.puffs.puffs.remove(puff)
        self.setupUI()

    def preSetupUI(self):
        self.groups=Groups(self)
        self.trash=Trash(self)
        self.autoGroupEvents(self.udc['radius'])
        self.setupUI()

    def setupUI(self):
        self.setWindowTitle('Puff Analyzer - {}'.format(self.data_window.name))
        self.setGeometry(QRect(360, 368, 1552, 351))
        
        
        
        if self.l is not None:
            QWidget().setLayout(self.layout())
        self.l=QGridLayout(self)
        self.setLayout(self.l)
        self.area = DockArea()
        self.l.addWidget(self.area)
        self.d1 = Dock("3D Fit", size=(600,300))
        self.d2 = Dock("Measured amplitude over time", size=(600,400))
        self.d3 = Dock("Control Panel", size=(100,400))
        self.d4 = Dock("Scatter Plot", size=(100,400))
        self.area.addDock(self.d1,'left')
        self.area.addDock(self.d2,'right')
        self.area.addDock(self.d3,'right')
        self.area.addDock(self.d4,'below',self.d1)
        
        first_puff = self.puffs[0]
        self.threeD_plot = threeD_plot(first_puff)
        self.d1.addWidget(self.threeD_plot)
        self.threeD_plot_ON = True
        self.trace_plot = pg.PlotWidget()
        if first_puff is not None:
            first_puff.plot(self.trace_plot)
        self.d2.addWidget(self.trace_plot)

        self.control_panel = QGridLayout()
        self.currentPuff_spinbox = QSpinBox()
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.currentPuff_spinbox.valueChanged.connect(self.setCurrPuff)
        self.discardButton = QPushButton('Delete Puff')
        self.discardButton.pressed.connect(self.discard_currPuff)
        self.togglePuffsButton = QPushButton('Toggle Puffs')
        self.togglePuffsButton.pressed.connect(self.togglePuffs)
        self.toggleGroupsButton = QPushButton('Toggle Groups')
        self.toggleGroupsButton.pressed.connect(self.toggleGroups)
        self.toggleTrashButton = QPushButton('Toggle Trash')
        self.toggleTrashButton.pressed.connect(self.toggleTrash)
        self.toggle3DButton = QPushButton('Toggle 3D Plot')
        self.toggle3DButton.pressed.connect(self.toggle3D)
        self.filterButton = QPushButton('Filter')
        self.filterButton.pressed.connect(self.openFilterGUI)
        self.widenButton = QPushButton('Widen all puff durations')
        self.widenButton.pressed.connect(self.widenPuffDurations)
        self.refitButton = QPushButton('Refit Gaussians')
        self.refitButton.pressed.connect(self.refitGaussians)
        self.compareWithManualButton = QPushButton('Compare with manual pts')
        self.compareWithManualButton.pressed.connect(self.compareWithManual)
        self.exportButton = QPushButton('Export to Excel')
        self.exportButton.pressed.connect(self.export_gui)
        self.savePointsButton = QPushButton('Save points')
        self.savePointsButton.pressed.connect(self.savePoints)
        self.saveButton = QPushButton('Save (.flika)')
        self.saveButton.pressed.connect(self.save)
        self.deleteToolButton = QPushButton('Delete puffs in Current ROI')
        self.deleteToolButton.pressed.connect(self.deleteTool)        
        
        self.control_panel.addWidget(self.currentPuff_spinbox,0,0)
        self.control_panel.addWidget(self.discardButton,1,0)
        self.control_panel.addWidget(self.togglePuffsButton,2,0)
        self.control_panel.addWidget(self.toggleGroupsButton,3,0)
        self.control_panel.addWidget(self.toggleTrashButton,4,0)
        self.control_panel.addWidget(self.widenButton,5,0)
        self.control_panel.addWidget(self.refitButton, 6, 0)
        self.control_panel.addWidget(self.filterButton,7,0)
        self.control_panel.addWidget(self.exportButton,8,0)
        self.control_panel.addWidget(self.toggle3DButton,9,0)
        self.control_panel.addWidget(self.compareWithManualButton, 10, 0)
        self.control_panel.addWidget(self.savePointsButton, 11, 0)
        self.control_panel.addWidget(self.saveButton, 12, 0)
        self.control_panel.addWidget(self.deleteToolButton, 13, 0)        
        self.control_panelWidget = QWidget()
        self.control_panelWidget.setLayout(self.control_panel)
        self.d3.addWidget(self.control_panelWidget)
        self.puffsVisible = True
        self.groupsVisible = False
        self.trashVisible = False
        self.currentPuff_spinbox.puffidx =- 1
        self.s1 = pg.ScatterPlotItem(size=5, pen=pg.mkPen(None)) #PUFFS
        self.data_window.imageview.addItem(self.s1)
        self.s1.sigClicked.connect(self.clicked)
        self.plot_puff_centroids(self.s1)


        def create_s2():
            '''I had to add these functions because pyqt was deleting them when they didn't have a parent.'''
            self.s2=pg.ScatterPlotItem(size=5,  brush=pg.mkBrush(0, 255, 0, 255),pen=pg.mkPen([0,0,0,255])) #GROUPS
            self.s2.sigClicked.connect(self.clickedGroup)
            self.s2.addPoints(pos=[group.pos for group in self.groups],data=self.groups)
            self.s2.destroyed.connect(create_s2)
        def create_s3():
            '''I had to add these functions because pyqt was deleting them when they didn't have a parent.'''
            self.s3=pg.ScatterPlotItem(size=5,  brush=pg.mkBrush(0, 255, 255, 255),pen=pg.mkPen([0,0,0,255])) #TRASH
            self.s3.sigClicked.connect(self.clickedTrash)
            for puff in self.trash:
                x=puff.kinetics['x']
                y=puff.kinetics['y']
                self.s3.addPoints(pos=[[x,y]],data=puff)
            self.s3.destroyed.connect(create_s3)
        create_s2()
        create_s3()
        
        self.s4 = pg.ScatterPlotItem(size=5, brush=pg.mkBrush(255, 255, 255, 120), pen=pg.mkPen([0, 0, 0, 255]))
        view = pg.GraphicsLayoutWidget()
        self.d4.addWidget(view)
        self.scatter_viewbox = ScatterViewBox()
        self.scatterPlot=view.addPlot(viewBox=self.scatter_viewbox)
        self.scatter_viewbox.drawFinishedSignal.connect(self.lasso_puffs)
        self.scatter_viewbox.EnterPressedSignal.connect(self.reset_scatter_colors)
        self.scatterPlot.addItem(self.s4)   
        self.lastClickedScatterPt = []
        self.s4.sigClicked.connect(self.clickedScatter)
        self.scatterPlot.axes['bottom']['item'].setLabel('Sigma of Puff Gaussian Fit (pixels)')
        self.scatterPlot.axes['left']['item'].setLabel('Amplitude of Puff Trace (F/F0)')
        self.groupAnalyzer = None
        self.lastClicked = None
        self.linkTif()
        self.updateScatter()
        self.setCurrPuff(0)
        self.show()

    def plot_puff_centroids(self, scatter):
        scatter.clear()
        for puff in self.puffs.puffs:
            x = puff.kinetics['x']
            y = puff.kinetics['y']
            scatter.addPoints(pos=[[x, y]], data=puff, brush=pg.mkBrush(puff.color), pen=pg.mkPen([0, 0, 0, 255]))

    def save(self):
        filename=self.data_window.filename
        filename=os.path.splitext(filename)[0]+'.flika'
        if len(filename) == 0:
            g.alert('The original data file has no filename, so the .flika file cannot be saved.')
            return None
        msg='Saving to {}'.format(filename)
        g.m.statusBar().showMessage(msg); print(msg)
        persistentInfo=PersistentInfo(self)
        
        with bz2.BZ2File(filename, 'w') as f:
            pickle.dump(persistentInfo, f)
        msg = 'Saved Flika file'
        g.alert(msg)

    def savePoints(self):
        g.m.statusBar().showMessage('Saving puff points'); print('Saving puff points')
        puffs=self.puffs.puffs
        pts=[]
        for puff in puffs:
            k=puff.kinetics
            pts.append(np.array([k['t_peak'], k['x'], k['y']]))
        pts=np.array(pts)
        filename=self.data_window.filename
        filename=os.path.splitext(filename)[0]+'_flika_pts.txt'
        np.savetxt(filename, pts)
        msg = 'Saved puff points'
        g.alert(msg)

    def clickedScatter(self, plot, points):
        p=points[0]
        starting_idx=p.data()
        curr_idx=[puff.starting_idx for puff in self.puffs].index(starting_idx)
        self.setCurrPuff(curr_idx)

    def updateScatter(self, X_axis='sigma', Y_axis='amplitude'):
        starting_indicies = []
        peak_amps = []
        stds = []
        times = []
        sigmas = []
        amps = []
        color = []
        results = dict()
        nEvents = len(self.puffs.puffs)
        for i in np.arange(nEvents):
            puff = self.puffs.puffs[i] # trashed puffs are not included in this list
            before = puff.kinetics['before']
            t0 = puff.kinetics['t_start']-before
            tf = puff.kinetics['t_end']-before
            trace_during = puff.trace[t0:tf+1]
            peak_amps.append(np.max(trace_during))
            amps.append(puff.kinetics['amplitude'])
            stds.append(np.std(trace_during))
            try:
                sigmas.append(puff.kinetics['sigma'])
            except KeyError:
                sigma = max([puff.kinetics['sigmax'], puff.kinetics['sigmay']])
                sigmas.append(sigma)
            color.append(puff.color)
            starting_indicies.append(puff.starting_idx)
        results['amplitude'] = np.array(amps)
        results['sigma'] = np.array(sigmas)
        
        self.s4.clear()
        n = len(starting_indicies)
        pos = np.array([results[X_axis],results[Y_axis]])
        spots = [{'pos': pos[:,i], 'data': starting_indicies[i], 'brush':pg.mkBrush(color[i])} for i in range(n)]
        self.s4.addPoints(spots)

    def lasso_puffs(self):
        pos_array=np.array([[self.s4.data[i][6], self.s4.data[i][0], self.s4.data[i][1]] for i in np.arange(len(self.s4.data))]) #array where each row is [starting_idx, x, y]
        self.s1.clear()
        for puff in self.puffs.puffs:
            x,y=pos_array[np.argwhere(pos_array[:,0]==puff.starting_idx)[0][0],1:]
            if self.scatter_viewbox.currentROI.contains(x,y):
                puff.color=self.scatter_viewbox.scatter_color
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s1.addPoints(pos=[[x,y]],data=puff, brush=pg.mkBrush(puff.color))
        self.updateScatter()

    def reset_scatter_colors(self):
        self.s1.clear()
        for puff in self.puffs.puffs:
            puff.color=(255,0,0,255)
            x=puff.kinetics['x']
            y=puff.kinetics['y']
            self.s1.addPoints(pos=[[x,y]],data=puff, brush=pg.mkBrush(puff.color), pen=pg.mkPen([0,0,0,255]))
        self.updateScatter()

    def toggle3D(self):
        if self.threeD_plot_ON:
            self.threeD_plot_ON=False
            self.d1.setParent(None)
            self.d1.label.setParent(None)
        else:
            self.threeD_plot_ON=True
            self.threeD_plot.update_puff(self.puffs.getPuff())
            self.d1 = Dock("3D Fit", size=(600,300))
            self.area.addDock(self.d1,'left')
            self.d1.addWidget(self.threeD_plot)

    def linkTif(self):
        self.clusterItem = pg.ImageItem(self.puffs.cluster_im[0])
        self.data_window.imageview.view.addItem(self.clusterItem)
        self.clusterItem.setOpacity(.5)
        self.clusterItem.setLevels(np.array([0, 1]))
        self.data_window.sigTimeChanged.connect(self.updateTime)        
        puff = self.puffs.getPuff()
        if puff is not None:
            x = np.floor(puff.kinetics['x']); y = np.floor(puff.kinetics['y'])
        else:
            x = 0
            y = 0
        roi_width = self.udc['roi_width']
        r = (roi_width-1)/2
        x0 = x - r
        x1 = x + r + 1
        y0 = y - r
        y1 = y + r + 1
        pts = [(x0, y0), (x0, y1), (x1, y1), (x1, y0), (x0, y0)]
        self.roi = makeROI('rectangle',pts,self.data_window, resizable=True)
        #self.data_window.deleteButtonSignal.disconnect(self.roi.deleteCurrentROI)
        self.redTraces = []
        self.data_window.keyPressSignal.connect(self.keyPressEvent)
        self.roi.plotSignal.connect(self.linkTracefig)
        self.roi.plot()

    def linkTracefig(self):
        g.currentTrace.finishedDrawingSignal.connect(self.drawRedOverlay) #hopefully this is not already connected
        g.currentTrace.partialThreadUpdatedSignal.connect(self.drawRedOverlay)
        g.currentTrace.p1.scene().sigMouseClicked.connect(self.clickedTrace)
        g.currentTrace.keyPressSignal.connect(self.keyPressEvent)
        self.drawRedOverlay()

    def updateTime(self,t):
        self.clusterItem.setImage(self.puffs.cluster_im[t], False) #If autolevels is not False, the clusterItem is invisible in a very small set of frames.

    def closeEvent(self, event):
        self.close()
        event.accept() # let the window close

    def close(self):
        if hasattr(self, 'roi'):
            if self.roi in self.data_window.rois:
                self.roi.delete()
            del self.roi
        if g.currentTrace is not None:
            for i in np.arange(len(self.redTraces)):
                g.currentTrace.p1.removeItem(self.redTraces[i][0])
            g.currentTrace.finishedDrawingSignal.disconnect(self.drawRedOverlay)
            g.currentTrace.p1.scene().sigMouseClicked.disconnect(self.clickedTrace)
            try:
                g.currentTrace.keyPressSignal.disconnect(self.keyPressEvent)
            except TypeError:
                pass
        if not self.data_window.closed:
            self.data_window.keyPressSignal.disconnect(self.keyPressEvent)
            self.data_window.sigTimeChanged.disconnect(self.updateTime) 
            self.data_window.imageview.view.removeItem(self.clusterItem)
            if self.data_window in g.windows:
                if self.groupsVisible:
                    self.data_window.imageview.removeItem(self.s2)
                if self.puffsVisible:
                    self.data_window.imageview.removeItem(self.s1)
            else:
                del self.data_window
        if self in g.windows:
            g.windows.remove(self)
        self.puffs = None

    def autoGroupEvents(self,radius=np.sqrt(2)):
        groups=self.groups[:]
        for group in groups:
            self.groups.remove(group) 
        puffs=self.puffs.puffs
        distances=np.zeros((len(puffs),len(puffs)))
        for i in np.arange(len(puffs)):
            x0,y0=(puffs[i].kinetics['x'], puffs[i].kinetics['y'])
            for j in np.arange(len(puffs)):
                x1,y1=(puffs[j].kinetics['x'], puffs[j].kinetics['y'])
                distances[i,j]=np.sqrt((x1-x0)**2+(y1-y0)**2)
        distances=distances<=radius
        groups=[]
        puffsAdded=set()
        for i in np.arange(len(puffs)):
            pfs=set([s[0] for s in np.argwhere(distances[i])])
            if len(pfs.intersection(puffsAdded))==0:
                groups.append(pfs)
                puffsAdded=puffsAdded.union(pfs)
            else:
                old_groups=[groups.index(group) for group in groups if len(group.intersection(pfs))>0]
                if len(old_groups)==1:
                    idx=old_groups[0]
                    groups[idx]=groups[idx].union(pfs)
                    puffsAdded=puffsAdded.union(pfs)
                else: #merge old groups and new puffs to them, delete old groups
                    groups.append(set.union(set.union(*[groups[i] for i in old_groups]),pfs))
                    for index in sorted(old_groups, reverse=True):
                        del groups[index]
        for group in groups:
            pfs=[puffs[idx] for idx in group]
            self.groups.append(Group(pfs))

    def compareWithManual(self):
        filename=g.settings['filename']
        try:
            directory=os.path.dirname(filename)
        except:
            directory=''
        prompt='Load points file containing manually selected points.'
        if filename is not None and directory != '':
            filename= QFileDialog.getOpenFileName(self, prompt, directory, '*.txt')
        else:
            filename= QFileDialog.getOpenFileName(self, prompt, '*.txt')
        filename=str(filename)
        simulated_puffs=np.loadtxt(filename)
        
    
    
        detected_puffs=self.puffs.puffs
        nearest_puffs=[]
        for puff in detected_puffs:
            k=puff.kinetics
            detected_puff=np.array([k['t_peak'],k['x'],k['y']])
            difference=np.sqrt(np.sum((simulated_puffs-detected_puff)**2,1))
            closest_idx=np.argmin(difference)
            #distance=difference[closest_idx]
            closest_puff=simulated_puffs[closest_idx]
            distance=detected_puff-closest_puff
            nearest_puffs.append([closest_idx,distance[0],distance[1],distance[2]])
            
        nearest_puffs=np.array(nearest_puffs)
        distances=np.sqrt(nearest_puffs[:,2]**2+nearest_puffs[:,3]**2)
        nTruePositives=np.count_nonzero(distances<2)
        print('Found {} puffs'.format(nTruePositives))
        
        #I need to finish this to include false positives and misses

    def toggleGroups(self):
        if self.groupsVisible:
            self.data_window.imageview.removeItem(self.s2)
        else:
            self.data_window.imageview.addItem(self.s2)
        self.groupsVisible=not self.groupsVisible

    def toggleTrash(self):
        if self.trashVisible:
            self.data_window.imageview.removeItem(self.s3)
        else:
            self.data_window.imageview.addItem(self.s3)
        self.trashVisible=not self.trashVisible

    def setCurrPuff(self, value, force=False):
        if force is False and self.currentPuff_spinbox.puffidx == value:
            return
        self.currentPuff_spinbox.puffidx = value
        self.puffs.setIndex(value)
        puff = self.puffs[value]
        if puff is None:
            return 
        if self.threeD_plot_ON:
            self.threeD_plot.update_puff(puff)
        self.trace_plot.clear()
        puff.plot(self.trace_plot)
        self.trace_plot.plotItem.autoRange()
        if self.roi.traceWindow is not None:
            rgnbounds = np.array(self.roi.traceWindow.region.getRegion())
            rgnbounds += puff.kinetics['t_peak']-int(np.mean(rgnbounds))
            self.roi.traceWindow.region.setRegion(tuple(rgnbounds))
#        
        if self.lastClicked is not None:
            self.lastClicked.resetPen()
            self.lastClicked.setBrush(self.lastClicked.data().color)
        point = [s for s in self.s1.points() if s.data() is puff][0]
        point.setPen('y', width=2)
        point.setBrush(puff.color)


        # move the roi to that point
        self.roi.center_around(puff.kinetics['x'], puff.kinetics['y'])
        #self.drawRedOverlay()

        self.lastClicked = point
        if self.currentPuff_spinbox.value() != self.puffs.index:
            self.currentPuff_spinbox.setValue(self.puffs.index)
        for p in self.lastClickedScatterPt:
            p.resetPen()
        p = [pt for pt in self.s4.points() if pt.data() == puff.starting_idx][0]
        if p is not None:
            p.setPen('y', width=2)
            self.lastClickedScatterPt = [p]

    def clicked(self, plot, points):
        point = points[0]
        puff = point.data()
        self.setCurrPuff(self.puffs.puffs.index(puff))

    def clickedGroup(self,plot,points):
        point=points[0]
        group=point.data()
        if self.groupAnalyzer is not None:
            self.groupAnalyzer=None
        self.groupAnalyzer=GroupAnalyzer(group)

    def clickedTrash(self,plot,points):
        point=points[0]
        puff=point.data()
        self.puffs.addPuffs([puff])
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        self.trash.removePuffs([puff])
        self.setCurrPuff(self.puffs.puffs.index(puff))

    def keyPressEvent(self,ev):
        nPuffs=len(self.puffs.puffs)
        old_idx=self.puffs.index
        if ev.key() == Qt.Key_Right:
            if old_idx==nPuffs-1:
                new_idx=0
            else:
                new_idx=old_idx+1
            self.setCurrPuff(new_idx)
        elif ev.key() == Qt.Key_Left:
            if old_idx==0:
                new_idx=nPuffs-1
            else:
                new_idx=old_idx-1
            self.setCurrPuff(new_idx)
        elif ev.key()==Qt.Key_Delete:
            if self.roi.mouseIsOver or self.roi.beingDragged:
                puffs=[pt.data() for pt in self.s1.points() if self.roi.contains_pts(pt.pos().x(), pt.pos().y())]
                self.discard_puffs(puffs)                
            else: 
                self.discard_currPuff()
        elif ev.key()==Qt.Key_G:
            #Group all events in roi
            groups=[pt.data() for pt in self.s2.points() if self.roi.contains_pts(pt.pos().x(), pt.pos().y())]
            if len(groups)==0:
                return
            puffs=list(itertools.chain(*[group.puffs for group in groups]))
            for group in groups:
                self.groups.remove(group) 
            self.groups.append(Group(puffs))
            self.s2.clear()
            for group in self.groups:
                self.s2.addPoints(pos=[group.pos],data=group)
        elif ev.key()==Qt.Key_U:
            groups=[pt.data() for pt in self.s2.points() if self.roi.contains_pts(pt.pos().x(), pt.pos().y())]
            if len(groups)==0:
                return
            puffs=list(itertools.chain(*[group.puffs for group in groups]))
            for group in groups:
                self.groups.remove(group) 
            self.groups.extend([Group([p]) for p in puffs])
            self.s2.clear()
            for group in self.groups:
                self.s2.addPoints(pos=[group.pos],data=group)

    def discard_currPuff(self):
        puff=self.puffs.getPuff()
        self.discard_puffs([puff])

    def discard_puffs(self,puffs):
        self.trash.addPuffs(puffs)
        self.groups.removePuffs(puffs)
        self.puffs.removePuffs(puffs)
        if len(self.puffs.puffs)==0:
            self.close()
            return 
        self.lastClickedScatterPt=[]
        self.updateScatter()
        self.setCurrPuff(self.puffs.index,force=True)
        self.currentPuff_spinbox.setMaximum(len(self.puffs.puffs)-1)
        # remove points from scatter plot
      

    def deleteTool(self):    
        #GD edit - added to allow deletion of multiple puffs at once
        puffs=[pt.data() for pt in self.s1.points() if self.data_window.currentROI.contains_pts(pt.pos().x(), pt.pos().y())]
        self.discard_puffs(puffs) 


    def togglePuffs(self):
        if self.puffsVisible:
            self.data_window.imageview.removeItem(self.s1)
        else:
            self.data_window.imageview.addItem(self.s1)
        self.puffsVisible=not self.puffsVisible

    def drawRedOverlay(self):
        puffs = [pt.data() for pt in self.s1.points() if self.roi.contains_pts(pt.pos().x(), pt.pos().y())]
        times = [[puff.kinetics['t_start'],puff.kinetics['t_end']+1] for puff in puffs]
        data = g.currentTrace.rois[g.currentTrace.get_roi_index(self.roi)]['p1trace'].getData()[1]
        x = np.array([np.arange(*times[i]) for i in np.arange(len(times))])
        traces = [data[time] for time in x]
        y = np.array(traces)
        for i in np.arange(len(self.redTraces)):
            g.currentTrace.p1.removeItem(self.redTraces[i][0])
        self.redTraces = []
        
        for i in np.arange(len(x)):
            pl = g.currentTrace.p1.plot(x[i],y[i],pen=pg.mkPen('r'))
            self.redTraces.append([pl,puffs[i]])
        currentPuff = self.puffs.getPuff()
        if currentPuff in puffs:
            idx = puffs.index(currentPuff)
            self.redTraces[idx][0].setPen(pg.mkPen(color='r',width=3))

    def clickedTrace(self,ev):
        self.EEEE=ev
        pos=ev.pos()
        pos=g.currentTrace.vb.mapSceneToView(pos)
        t=pos.x()
        puffs=[pt.data() for pt in self.s1.points() if self.roi.contains_pts(pt.pos().x(), pt.pos().y())]
        times=[[puff.kinetics['t_start'],puff.kinetics['t_end']] for puff in puffs]
        try:
            index=[(t<=time[1] and t>=time[0]) for time in times].index(True)
        except ValueError:
            return
        puff=puffs[index]
        if ev.button()==1:
            self.setCurrPuff(self.puffs.puffs.index(puff))

    def openFilterGUI(self):
        print("This will eventually allow you to filter puffs by things like amplitude and duration, but I haven't implemented it yet")

    def widenPuffDurations(self):
        puffs = self.puffs.puffs
        mt = len(self.data_window.image)
        for puff in puffs:
            [(t0, t1), _, _] = puff.bounds
            t0 -= 1
            if t0 < 0:
                t0 = 0
            t1 += 1
            if t1 >= mt:
                t1 = mt-1
            puff.bounds[0] = (t0,t1)
            puff.kinetics = puff.calcRiseFallTimes(puff.kinetics)
        puff = self.puffs.getPuff()
        self.trace_plot.clear()
        puff.plot(self.trace_plot)
        self.updateScatter()
        self.drawRedOverlay()

    def refitGaussians(self):
        self.puffs.refit_gaussians()
        puff = self.puffs.getPuff()
        self.updateScatter()
        self.drawRedOverlay()
        self.plot_puff_centroids(self.s1)
        curr_puff_idx = self.puffs.puffs.index(puff)
        self.lastClicked = None
        self.setCurrPuff(curr_puff_idx, force=True)

    def export_gui(self):
        filename = g.settings['filename']
        directory = os.path.dirname(filename)
        if filename is not None:
            filename = save_file_gui('Export Puff Info', directory, '*.xlsx')
        else:
            filename = save_file_gui('Export Puff Info', '', '*.xlsx')
        if filename == '' or filename is None:
            return False
        else:
            self.export(filename)

    def export(self, filename):
        ''' This function saves out all the info about the puffs in an Excel spreadsheet.      
        '''
        import openpyxl
        if StrictVersion(openpyxl.__version__) < StrictVersion('2.5.2'):
            g.alert('You are using openpyxl version {}. Upgrade to at least version 2.5.2 to export Excel file.'.format(openpyxl.__version__))
            return None
        from openpyxl import Workbook
        g.m.statusBar().showMessage('Saving {}'.format(os.path.basename(filename)))

        workbook = Workbook() 
        sheet = workbook.create_sheet(title='Event Data')
        header=['Group #','Group x','Group y','No. Events','Max Amp','x','y','t_peak','Amplitude','Integral','sigmax','sigmay','angle','r20','r50','r80','r100','f80','f50','f20','f0']
        for j in np.arange(len(header)):
            sheet.cell(row=1, column=j+1).value = header[j]
        row = 2
        groupN = 1
        for group in self.groups:
            groupx, groupy = group.pos
            nEvents = len(group.puffs)
            maxAmp = np.max([puff.kinetics['amplitude'] for puff in group.puffs])
            sheet.cell(row=row, column=1).value = groupN
            sheet.cell(row=row, column=2).value = groupx
            sheet.cell(row=row, column=3).value = groupy
            sheet.cell(row=row, column=4).value = nEvents
            sheet.cell(row=row, column=5).value = maxAmp
            for puff in group.puffs:
                r = str(row)
                k = puff.kinetics
                sheet.cell(row=row, column=6).value = k['x']
                sheet.cell(row=row, column=7).value = k['y']
                try: sheet.cell(row=row, column=8).value = int(k['t_peak'])
                except ValueError: pass
                sheet.cell(row=row, column=9).value = k['amplitude']
                sheet.cell(row=row, column=10).value = k['integral']                
                
                if 'sigmax' in k.keys():
                    sheet.cell(row=row, column=11).value = k['sigmax']
                    sheet.cell(row=row, column=12).value = k['sigmay']
                    sheet.cell(row=row, column=13).value = k['angle']
                else:
                    sheet.cell(row=row, column=11).value = k['sigma']
                try: sheet.cell(row=row, column=14).value = int(k['r20'])
                except ValueError: pass
                try: sheet.cell(row=row, column=15).value = int(k['r50'])
                except ValueError: pass
                try: sheet.cell(row=row, column=16).value = int(k['r80'])
                except ValueError: pass
                try: sheet.cell(row=row, column=17).value = int(k['t_peak']-k['t_start'])
                except ValueError: pass
                try: sheet.cell(row=row, column=18).value = int(k['f80'])
                except ValueError: pass
                try: sheet.cell(row=row, column=19).value = int(k['f50'])
                except ValueError: pass
                try: sheet.cell(row=row, column=20).value = int(k['f20'])
                except ValueError: pass
                try: sheet.cell(row=row, column=21).value = int(k['f0'])
                except ValueError: pass
                row += 1
            groupN += 1
            
        sheet = workbook.create_sheet(title="Group traces")
        groupN = 1
        for group in self.groups:
            x = int(np.floor(group.pos[0]))
            y = int(np.floor(group.pos[1]))
            roi_width = g.puffAnalyzer.udc['roi_width']
            r = (roi_width-1)/2
            xmin = int(max(x - r, 0))
            ymin = int(max(y - r, 0))
            xmax = int(min(x + r, self.data_window.mx))
            ymax = int(min(y + r, self.data_window.my))
            trace = self.data_window.image[:, xmin:xmax+1, ymin:ymax+1]
            trace = np.mean(np.mean(trace, 1), 1)
            sheet.cell(row=1, column=groupN).value = "Group #{}".format(groupN)
            if trace.dtype == np.float16:
                trace = trace.astype(np.float)
            for i in np.arange(len(trace)):
                sheet.cell(row=i+2, column=groupN).value = trace[i]
            groupN += 1
        
        sheet = workbook.create_sheet(title="Peak aligned Event Traces")
        groupN = 1
        max_peak_idx = np.max([puff.kinetics['t_peak']-puff.kinetics['t_start'] for puff in self.puffs])
        for puff in self.puffs.puffs:

            peak_idx = puff.kinetics['t_peak']-puff.kinetics['t_start']
            trace = puff.trace
            if trace.dtype == np.float16:
                trace = trace.astype(np.float)
            for i in np.arange(len(trace)):
                offset = max_peak_idx-peak_idx
                sheet.cell(row=offset+i+1, column=groupN).value = trace[i]
            groupN += 1
        workbook.remove(workbook.worksheets[0])  # Get rid of blank first worksheet.
        workbook.save(filename)
        g.alert('Successfully saved {}'.format(os.path.basename(filename)))
        #g.m.statusBar().showMessage('Successfully saved {}'.format(os.path.basename(filename)))
            

class threeD_plot(gl.GLViewWidget):
    def __init__(self,puff,parent=None):
        super(threeD_plot,self).__init__(parent)
        self.setCameraPosition(distance=150,elevation=30,azimuth=90)
        if puff is not None:
            image=np.copy(puff.gaussianFit)
            mmax=np.max(puff.mean_image)
            image=image/mmax
        else:
            image=np.zeros((10,10))
        self.p1 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        ##    red   = pow(z * colorMap[0] + colorMap[1], colorMap[2])
        ##    green = pow(z * colorMap[3] + colorMap[4], colorMap[5])
        ##    blue  = pow(z * colorMap[6] + colorMap[7], colorMap[8])
        self.p1.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p1.scale(1, 1, 15.0)
        if puff is not None:
            self.p1.translate(-puff.udc['paddingXY'], -puff.udc['paddingXY'], 0)
        self.addItem(self.p1)
        
        if puff is not None:
            image=np.copy(puff.mean_image)
            image=image/mmax
        else:
            image=np.zeros((10,10))
        self.p2 = gl.GLSurfacePlotItem(z=image, shader='heightColor')
        self.p2.shader()['colorMap'] = np.array([1, 0, 1, 1, .3, 2, 1, .4, 1])
        self.p2.scale(1, 1, 15.0)
        if puff is not None:
            self.p2.translate(-puff.udc['paddingXY'], -puff.udc['paddingXY'], 0)
        self.addItem(self.p2)
        
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
    def update_puff(self,puff):
        self.p1.translate(-self.shiftx,0,0)
        self.p2.translate(self.shiftx,0,0)
        if puff is None:
            return
        image=np.copy(puff.gaussianFit)
        mmax=np.max(puff.mean_image)
        image=image/mmax
        self.p1.setData(z=image)
        image=np.copy(puff.mean_image)
        image=image/mmax
        self.p2.setData(z=image)
        self.shiftx=int(np.ceil(image.shape[0]/2))
        self.p1.translate(self.shiftx,0,0)
        self.p2.translate(-self.shiftx,0,0)
        #self.setMinimumHeight(300)




class Trash(list):
    def __init__(self,puffAnalyzer):
        super(Trash,self).__init__()
        self.puffAnalyzer=puffAnalyzer
    def addPuffs(self,puffs):
        self.extend(puffs)
        pos=[[puff.kinetics['x'],puff.kinetics['y']] for puff in puffs]
        #self.puffAnalyzer.s3.addPoints(pos=pos,data=puffs)
        scatterAddPoints(self.puffAnalyzer.s3,pos,puffs)
    def removePuffs(self,puffs):
        for puff in puffs:
            if puff in self:
                self.remove(puff)
        print('removing puffs, replacing trash scatterplot')
        s=self.puffAnalyzer.s3
        idxs=[]
        for puff in puffs:
            idxs.append([point['data'] for point in s.data].index(puff))
        scatterRemovePoints(self.puffAnalyzer.s3,idxs)


def findextrema(v):
    v_prime=np.concatenate(([0], np.diff(v)))
    zero_crossing=np.concatenate((np.diff(np.sign(v_prime)),[0]))
    maxima=[1 if v < -1 else 0 for v in zero_crossing]
    minima=[1 if v > 1 else 0 for v in zero_crossing]
    if np.sign(v_prime[-1])>0:
        maxima[-1]=1
    else:
        minima[-1]=1;
    if np.sign(v_prime[0])>0:
        maxima[0]=1
    else:
        minima[0]=1
    maxima=np.nonzero(maxima)[0]
    minima=np.nonzero(minima)[0]
    return maxima, minima


def scatterRemovePoints(scatterplot,idxs):
    i2=[i for i in np.arange(len(scatterplot.data)) if i not in idxs]
    points=scatterplot.points()
    points=points[i2]
    spots=[{'pos':points[i].pos(),'data':points[i].data(),'brush':points[i].brush()} for i in np.arange(len(points))]
    scatterplot.clear()
    scatterplot.addPoints(spots)


def scatterAddPoints(scatterplot,pos,data):
    points=scatterplot.points()
    spots=[{'pos':points[i].pos(),'data':points[i].data()} for i in np.arange(len(points))]
    spots.extend([{'pos':pos[i],'data':data[i]} for i in np.arange(len(pos))])
    scatterplot.clear()
    scatterplot.addPoints(spots)


class ScatterViewBox(ClusterViewBox):
    def __init__(self, *args, **kwds):
        ClusterViewBox.__init__(self, *args, **kwds)
        self.colorDialog=QColorDialog()
        self.colorDialog.colorSelected.connect(self.colorSelected)
        self.scatter_color=(255,239,5,255)
    def keyPressEvent(self,ev):
        if ev.key() == Qt.Key_Enter or ev.key() == Qt.Key_Return:
            self.EnterPressedSignal.emit()
    def mouseDragEvent(self, ev):
        if ev.button() == Qt.RightButton:
            ev.accept()
            self.ev=ev
            if ev.isStart():
                pt=self.mapSceneToView(self.ev.buttonDownScenePos())
                self.x=pt.x() # this sets x and y to the button down position, not the current position
                self.y=pt.y()
                #print("Drag start x={},y={}".format(self.x,self.y))
                if self.currentROI is not None:
                    self.currentROI.delete()
                self.currentROI=ROI(self,self.x,self.y)
            if ev.isFinish():
                self.currentROI.drawFinished()
                self.drawFinishedSignal.emit()
            else: # if we are in the middle of the drag between starting and finishing
                pt=self.mapSceneToView(self.ev.scenePos())
                self.x=pt.x() # this sets x and y to the button down position, not the current position
                self.y=pt.y()
                #print("Drag continuing x={},y={}".format(self.x,self.y))
                self.currentROI.extend(self.x,self.y)
        else:
            pg.ViewBox.mouseDragEvent(self, ev)
    def mouseClickEvent(self,ev):
        if ev.button() == Qt.RightButton:
            ev.accept()
            self.colorDialog.open() #chage the color
        else:
            pg.ViewBox.mouseClickEvent(self, ev)
    def colorSelected(self, color):
        self.color=color
        if color.isValid():
            self.scatter_color=tuple((np.array(self.color.getRgbF())*255).astype(np.int))
        

class PersistentInfo:
    def __init__(self,puffAnalyzer):
        self.clusters=puffAnalyzer.clusters.clusters # list of arrays of idxs for each cluster
        self.pixel_idxs=puffAnalyzer.clusters.idxs #[t,x,y] for every pixel determined to be in a puff
        self.puffs={puff.starting_idx:
                         {'kinetics':puff.kinetics, 
                         'trace':puff.trace, 
                         'gaussianParams':puff.gaussianParams, 
                         'mean_image':puff.mean_image,
                         'gaussianFit':puff.gaussianFit,
                         'color':puff.color,
                         'trashed':False} for puff in puffAnalyzer.puffs.puffs}
        self.puffs.update({puff.starting_idx:
                        {'kinetics':puff.kinetics, 
                         'trace':puff.trace, 
                         'gaussianParams':puff.gaussianParams, 
                         'mean_image':puff.mean_image,
                         'gaussianFit':puff.gaussianFit,
                         'color':puff.color,
                         'trashed':True}  for puff in puffAnalyzer.trash})
        self.groups=[[puff.starting_idx for puff in group.puffs] for group in puffAnalyzer.groups]
        self.udc=puffAnalyzer.udc
        self.movieShape=puffAnalyzer.clusters.movieShape
        self.data_window_commands=puffAnalyzer.data_window.commands


def run_demo():
    from .puff_simulator.puff_simulator import simulate_puffs # from plugins.detect_puffs.puff_simulator.puff_simulator import simulate_puffs
    simulate_puffs(nFrames=1000, nPuffs=20)
    baseline = -5  # This is the mean pixel value in the absence of photons.
    subtract(baseline)
    data_window = ratio(0, 30, 'average')  # ratio(first_frame, nFrames, ratio_type). Now we are in F/F0
    data_window.setName('Data Window (F/F0)')
    norm_image = data_window.image - 1
    norm_window = Window(norm_image)
    norm_window.setName('Normalized Window')
    blurred_window = gaussian_blur(2, norm_edges=True, keepSourceWindow=True)
    blurred_window.setName('Blurred Window')
    threshold_cluster(data_window, blurred_window, blurred_window, blur_thresh=.1)
